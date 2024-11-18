from openpyxl import load_workbook
import datetime

sk_pathf1='C:/Users/Administrator/Desktop/'
sk_pathf2='C:/Users/Administrator/Desktop/'
sk_pathf3='C:/Users/Administrator/Desktop/'
sk_pathf4='C:/Users/Administrator/Desktop/'
path1=r'C:\Users\Administrator\Desktop'

def parse_date(x):
        if isinstance(x, str):
            try:
                return datetime.datetime.strptime(x, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return datetime.datetime.min  # 如果无法转换字符串，返回最小日期时间
        elif isinstance(x, datetime.datetime):
            return x
        else:
            return datetime.datetime.min  # 非日期时间字符串的处理方式

def sort_data(path,temp):
    # 加载Excel文件
    file_path = path+temp+'.xlsx'
    wb = load_workbook(filename=file_path)

    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 假设第一列是日期列
        data_rows = list(ws.iter_rows(values_only=True))
        header, data = data_rows[0], data_rows[1:]

        # 按第一列（日期）排序，处理None值
        #
        sorted_data = sorted(data, key=lambda x: parse_date(x[0]) if x[0] is not None else datetime.datetime.min)
    


        # 清除旧数据并写入新数据
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None

        # 写入排序后的数据
        for row_idx, row_data in enumerate(sorted_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # 保存到新文件，以保留原文件不变
    sorted_file_path = file_path
    wb.save(filename=sorted_file_path)

def sk1_sort(temp):
    sort_data(sk_pathf1,temp)

def sk2_sort(temp):
    sort_data(sk_pathf2,temp)
    
def sk3_sort(temp):
    sort_data(sk_pathf3,temp)

def sk4_sort(temp):
    sort_data(sk_pathf4,temp)

def normal_sort(temp):
    sort_data(path1,temp)




